import pandas as pd
import pingouin as pg
import sys
import seaborn as sns
import matplotlib.pyplot as plt
import os
import openpyxl



'''df = pd.DataFrame({'Q1': [1, 2, 2, 3, 2, 2, 3, 3, 2, 3],
                   'Q2': [1, 1, 1, 2, 3, 3, 2, 3, 3, 3],
                   'Q3': [1, 1, 2, 1, 2, 3, 3, 3, 2, 3]})

print(df)
temp=pg.cronbach_alpha(data=df)
print("\n")
print("Cronbach's_alpha_score=",temp,"\n")'''

def cronsbach_alpha(data,prefix):
    print(data,"\n")
    print("Correlation \n \n",data.corr(),"\n")
    sns.heatmap(data.corr(), cmap='RdBu',vmin=-1, vmax=1, linewidth=0.5, annot=True,annot_kws={'fontsize':11, 'fontweight':'bold'},square=True)
    plt.title(prefix+"_correlation")
    plt.show()
    df_trans=data.transpose()
    # Drop the column Qusetionarie => value is not neumeric
    df_trans=df_trans.drop(labels='Qusetionarie',  axis=0)
    print(df_trans)
    df_dict = df_trans.to_dict()
    print("\n\n",df_dict,"\n")
    df_2= pd.DataFrame(df_dict)
    temp_2=pg.cronbach_alpha(data=df_2,ci=.95)
    print("\n")
    print( prefix+"_Cronbach's_alpha_score=",temp_2,"\n")
    pass

def main():
    file_path=sys.argv[1]
    sheet_no=sys.argv[2]
    data = pd.read_excel(file_path,sheet_name= int(sheet_no))
    wb = openpyxl.load_workbook(file_path)
    prefix=wb.sheetnames[int(sheet_no)]
    cronsbach_alpha(data,prefix)
    #print(wb.sheetnames)
    pass

main()
